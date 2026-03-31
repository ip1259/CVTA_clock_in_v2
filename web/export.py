import os
from io import BytesIO

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date, time, timedelta
import calendar
import tempfile

from sqlalchemy import extract
from database.database_manager import db_manager
from database.base import Employee, Record, Holiday, ShiftAssignment, Schedule


def _change_cell_color(cell: Cell, color: str) -> None:
    """
    改變指定儲存格的顏色。

    Args:
        cell (Cell): 儲存格物件。
        color (str): 欲設定的顏色，可為"green" "red" "yellow" "gray"。

    Returns:
        None
    """

    fill = PatternFill(fill_type="solid")
    color = color.lower()
    match color:
        case "green":
            fill.start_color = "CCFFCC"
            fill.end_color = "CCFFCC"
        case "red":
            fill.start_color = "FF9980"
            fill.end_color = "FF9980"
        case "yellow":
            fill.start_color = "FFEB9C"
            fill.end_color = "FFEB9C"
        case "gray":
            fill.start_color = "F2F2F2"
            fill.end_color = "F2F2F2"
        case _:
            raise ValueError(
                "Invalid color. Please specify 'green','red','yellow' or 'gray'.")

    cell.fill = fill


def _convert_to_daily_records(employee_id: int, query_year: int, query_month: int) -> list[dict]:
    """
    將打卡紀錄轉換為每日紀錄

    Args:
        employee_id: 查詢員工 ID
        query_year: 查詢年份
        query_month: 查詢月份

    Returns:
        list[dict]: 每日紀錄列表，每個元素包含姓名、日期、星期幾、開始時間和結束時間
    """
    session = db_manager.get_session(query_year)
    weekday_map = {0: "一", 1: "二", 2: "三", 3: "四", 4: "五", 5: "六", 6: "日"}

    # 1. 取得員工資訊與預設班表
    emp = session.query(Employee).get(employee_id)
    if not emp:
        session.close()
        return []
    employee_name = emp.name

    # 2. 獲取該月打卡紀錄 (擴大範圍至前後一天，以處理跨月跨日班次)
    num_days = calendar.monthrange(query_year, query_month)[1]
    first_day_dt = datetime(query_year, query_month, 1)
    last_day_dt = datetime(query_year, query_month, num_days) + timedelta(days=1)

    records = session.query(Record).filter(
        Record.employee_id == employee_id,
        Record.record_time >= first_day_dt - timedelta(days=1),
        Record.record_time <= last_day_dt + timedelta(days=1)
    ).all()

    # 3. 準備當月每一天的初始資料結構
    daily_records = {}

    for day in range(1, num_days + 1):
        current_date = date(query_year, query_month, day)

        # 判定當天是否需要上班
        # A. 檢查是否有特殊節假日/補班定義
        holiday_def = session.query(Holiday).filter(
            Holiday.date == current_date).first()
        shiftAssignment_def = session.query(ShiftAssignment).filter(
            ShiftAssignment.date == current_date,
            ShiftAssignment.employee_id == employee_id
        ).first()

        if holiday_def:
            is_workday = holiday_def.is_workday
            note = holiday_def.description
        elif shiftAssignment_def:
            is_workday = True
            note = ""
        else:
            is_workday = current_date.weekday() < 5  # 平日預設上班
            note = ""

        # B. 取得當日適用的班表 (排班表優先於預設班表)
        assigned_shift = session.query(ShiftAssignment).filter(
            ShiftAssignment.employee_id == emp.id,
            ShiftAssignment.date == current_date
        ).first()

        current_schedule = assigned_shift.schedule if assigned_shift else emp.schedule

        daily_records[current_date] = {
            "name": employee_name,
            "date": current_date.strftime("%Y/%m/%d"),
            "weekday": weekday_map[current_date.weekday()],
            "start_time": None,
            "end_time": None,
            "is_workday": is_workday,
            "note": note,
            # 儲存班表標準時間用於後續比對
            "std_start": current_schedule.job_start if current_schedule else None,
            "std_end": current_schedule.job_end if current_schedule else None
        }

    # 4. 填充打卡數據 (按班次時間區間匹配，而非單純按日期)
    for d_date, day_item in daily_records.items():
        s_time = day_item["std_start"]
        e_time = day_item["std_end"]
        if not s_time or not e_time:
            continue

        # 計算該班次的起始與結束邊界
        start_dt = datetime.combine(d_date, s_time)
        if e_time < s_time:  # 跨日班
            end_dt = datetime.combine(d_date + timedelta(days=1), e_time)
        else:
            end_dt = datetime.combine(d_date, e_time)

        # 搜尋區間：上班前 4 小時到下班後 4 小時 (可視需求調整緩衝長度)
        search_start = start_dt - timedelta(hours=4)
        search_end = end_dt + timedelta(hours=4)

        # 找出屬於此班次的打卡紀錄
        shift_punches = [r.record_time for r in records if search_start <= r.record_time <= search_end]

        if shift_punches:
            p_min = min(shift_punches)
            p_max = max(shift_punches)
            day_item["start_time"] = p_min.time()
            day_item["start_dt"] = p_min  # 暫存完整的 datetime 以利後續計算
            day_item["end_time"] = p_max.time()
            day_item["end_dt"] = p_max    # 暫存完整的 datetime 以利後續計算

    # 5. 清洗數據 (處理單次打卡邏輯與格式化)
    result = []
    for d_date, r in daily_records.items():
        # 如果當天只有一筆打卡，且間隔過短 (邏輯同原版)
        if r["start_time"] and r["end_time"] and r["std_start"]:
            diff = r["end_dt"] - r["start_dt"]
            if diff < timedelta(hours=1):
                # 重新計算班次中點以判定這唯一的一筆是簽到還是簽退
                s_dt = datetime.combine(d_date, r["std_start"])
                if r["std_end"] < r["std_start"]:
                    e_dt = datetime.combine(d_date + timedelta(days=1), r["std_end"])
                else:
                    e_dt = datetime.combine(d_date, r["std_end"])
                mid_dt = s_dt + (e_dt - s_dt) / 2

                if r["start_dt"] <= mid_dt:
                    r["end_time"] = None
                else:
                    r["start_time"] = None

        # 格式化輸出
        r["start_time"] = r["start_time"].strftime(
            "%H:%M") if r["start_time"] else ""
        r["end_time"] = r["end_time"].strftime(
            "%H:%M") if r["end_time"] else ""

        # 移除不需要回傳給 Excel 的內部欄位
        cleaned_item = {
            "name": r["name"], "date": r["date"], "weekday": r["weekday"],
            "start_time": r["start_time"], "end_time": r["end_time"],
            "note": r["note"], "is_workday": r["is_workday"],
            "std_start": r["std_start"], "std_end": r["std_end"]
        }
        result.append(cleaned_item)

    session.close()
    return result


def _create_attendance_workbook(employee_ids: list[int], year: int, month: int):
    """
    建立考勤日報表 Excel 工作簿

    Args:
      employee_ids: 員工 ID 列表
      year: 年份
      month: 月份

    Returns:
      openpyxl.Workbook: 建立好的 Excel 工作簿
    """
    session = db_manager.get_session()
    # 建立新的工作簿
    wb = Workbook()
    border_style = Side(border_style="thin", color="FF000000")
    border = Border(left=border_style, right=border_style,
                    top=border_style, bottom=border_style)

    for i, emp_id in enumerate(employee_ids):
        emp = session.query(Employee).get(emp_id)
        if not emp:
            continue

        # 使用 "姓名(ID)" 作為分頁名稱，避免同名衝突
        sheet_name = f"{emp.name}({emp.id})"
        if i == 0:
            sheet = wb.active
        else:
            sheet = wb.create_sheet()
        sheet.title = sheet_name

        # 合併第一行並設定標題格式
        sheet.merge_cells('A1:F1')
        title_cell: Cell = sheet['A1']
        title_cell.value = f"{year}年{month}月 考勤日報表"
        title_font = Font(name='微軟正黑體', size=16, bold=True)
        title_cell.font = title_font
        title_cell.alignment = Alignment(
            horizontal='center', vertical='center')
        title_cell.border = border

        # 設定第二行為欄標題
        headers = ['姓名', '打卡日期', '星期', '簽到時間', '簽退時間', '備註']
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(name='微軟正黑體', size=12, bold=True)
            sheet.column_dimensions[get_column_letter(col)].width = "13.97"
            cell.border = border

        # 獲取並填入資料
        daily_data = _convert_to_daily_records(emp_id, year, month)
        for r_idx, row_data in enumerate(daily_data, start=3):
            # 將 dict 轉成 list 填入 (過濾掉內部使用的 flag)
            values = [row_data['name'], row_data['date'], row_data['weekday'],
                      row_data['start_time'], row_data['end_time'], row_data['note']]

            for c_idx, val in enumerate(values, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = val
                cell.border = border

                is_workday = row_data['is_workday']
                std_start = row_data['std_start'].strftime(
                    "%H:%M") if row_data['std_start'] else "00:00"
                std_end = row_data['std_end'].strftime(
                    "%H:%M") if row_data['std_end'] else "23:59"

                # 著色邏輯
                if cell.column in [4, 5]:
                    if not is_workday:
                        _change_cell_color(cell, "gray")
                    elif val == "":
                        _change_cell_color(cell, "yellow")
                    if cell.column == 4:
                        if is_workday and val != "" and val > std_start:
                            _change_cell_color(cell, "red")
                    elif cell.column == 5:
                        if is_workday and val != "" and val < std_end:
                            _change_cell_color(cell, "red")

    session.close()
    return wb


def export_attendance(wb: Workbook, year: int, month: int, export_type: str = "file"):
    """
    將 Workbook 物件匯出

    :param export_type: 'file' 回傳暫存路徑, 'bytes' 回傳 BytesIO 物件
    """
    filename = f"{year}-{month} 考勤日報表.xlsx"

    if export_type == "bytes":
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    else:
        save_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(save_path)
        return save_path
